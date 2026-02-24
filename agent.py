

from __future__ import annotations

import copy
import datetime
import re
import tempfile
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from google.adk.agents import Agent
# from google.adk.models.lite_llm import LiteLlm
from openpyxl import Workbook, load_workbook

# Preferred column orderings to keep XLSForm sheets tidy.
_PREFERRED_COLUMN_ORDER: Dict[str, List[str]] = {
    "survey": [
        "type",
        "name",
        "label",
        "hint",
        "required",
        "relevant",
        "appearance",
        "default",
        "constraint",
        "constraint_message",
        "calculation",
        "choice_filter",
        "repeat_count",
        "autoplay",
        "image",
        "audio",
        "video",
    ],
    "choices": [
        "list_name",
        "name",
        "label",
        "choice_filter",
    ],
    "settings": [
        "form_title",
        "form_id",
        "default_language",
        "version",
        "public_key",
        "submission_url",
        "instance_name",
    ],
}

_TEMPLATE_URL = (
    "https://github.com/getodk/xlsform-template/raw/main/ODK%20XLSForm%20Template.xlsx"
)
_TEMPLATE_CACHE_PATH = Path(tempfile.gettempdir()) / "odk_xlsform_template.xlsx"

# Common language names mapped to their BCP 47 / IANA subtags.
_CANONICAL_LANGUAGE_CODES: Dict[str, str] = {
    "Arabic": "ar",
    "Bengali": "bn",
    "Bulgarian": "bg",
    "Chinese": "zh",
    "Croatian": "hr",
    "Czech": "cs",
    "Danish": "da",
    "Dutch": "nl",
    "English": "en",
    "Finnish": "fi",
    "French": "fr",
    "German": "de",
    "Greek": "el",
    "Hebrew": "he",
    "Hindi": "hi",
    "Hungarian": "hu",
    "Indonesian": "id",
    "Italian": "it",
    "Japanese": "ja",
    "Korean": "ko",
    "Malay": "ms",
    "Norwegian": "no",
    "Persian": "fa",
    "Polish": "pl",
    "Portuguese": "pt",
    "Romanian": "ro",
    "Russian": "ru",
    "Serbian": "sr",
    "Somali": "so",
    "Spanish": "es",
    "Swahili": "sw",
    "Swedish": "sv",
    "Thai": "th",
    "Turkish": "tr",
    "Ukrainian": "uk",
    "Urdu": "ur",
    "Vietnamese": "vi",
}
_LANGUAGE_CODE_MAP: Dict[str, str] = {name.lower(): code for name, code in _CANONICAL_LANGUAGE_CODES.items()}
_LANGUAGE_CODE_MAP.update(
    {
        "farsi": "fa",
        "filipino": "fil",
        "kiswahili": "sw",
        "mandarin": "zh",
        "pashto": "ps",
        "swahili": "sw",
        "tagalog": "tl",
    }
)
_CODE_TO_NAME: Dict[str, str] = {}
for name, code in _CANONICAL_LANGUAGE_CODES.items():
    _CODE_TO_NAME.setdefault(code, name)


def _get_template_path() -> Optional[Path]:
    """Download the official ODK XLSForm template, caching it in the system temp dir."""
    if _TEMPLATE_CACHE_PATH.exists():
        return _TEMPLATE_CACHE_PATH
    try:
        urllib.request.urlretrieve(_TEMPLATE_URL, _TEMPLATE_CACHE_PATH)
        return _TEMPLATE_CACHE_PATH
    except Exception:
        return None


def _safe_form_id(title: str) -> str:
    """Convert an arbitrary title into a safe XLSForm id."""
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in title).strip("_")
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return cleaned or "odk_form"


def _normalize_language_tag(language: str) -> Dict[str, str]:
    """Return name/code/header fields for an input language string."""
    lang = (language or "").strip()
    if not lang:
        return {"header": "", "code": "", "name": ""}

    paren_match = re.match(r"^(.*?)(?:\s*\(([^)]+)\))\s*$", lang)
    if paren_match:
        name, code = paren_match.groups()
        name = name.strip()
        code = code.strip()
        normalized_code = code.lower()
        name = name or _CODE_TO_NAME.get(normalized_code, normalized_code)
        return {"header": f"{name}({normalized_code})", "code": normalized_code, "name": name}

    key = lang.lower()
    if key in _LANGUAGE_CODE_MAP:
        code = _LANGUAGE_CODE_MAP[key]
        name = _CODE_TO_NAME.get(code, lang.strip() or code)
        return {"header": f"{name}({code})", "code": code, "name": name}

    if re.fullmatch(r"[a-zA-Z]{2,3}(?:-[a-zA-Z0-9]{2,8})*", lang):
        code = lang.lower()
        name = _CODE_TO_NAME.get(code, lang)
        return {"header": f"{name}({code})", "code": code, "name": name}

    safe_code = "".join(ch.lower() for ch in lang if ch.isalnum() or ch == "-") or "und"
    name = lang
    return {"header": f"{name}({safe_code})", "code": safe_code, "name": name}


def _normalize_languages(languages: Sequence[str]) -> List[Dict[str, str]]:
    """Normalize a list of language names to include IANA codes."""
    normalized: List[Dict[str, str]] = []
    seen_headers: set[str] = set()
    for lang in languages or []:
        entry = _normalize_language_tag(lang)
        header = entry.get("header")
        if header and header not in seen_headers:
            seen_headers.add(header)
            normalized.append(entry)
    return normalized


def _language_headers_from_columns(columns: Sequence[str]) -> List[str]:
    """Extract unique language headers from language-bearing columns."""
    headers: List[str] = []
    for col in columns or []:
        m = re.match(r"^(label|hint|constraint_message|required_message|guidance_hint|image|audio|video)::(.+)$", str(col) or "")
        if m:
            header = m.group(2).strip()
            if header and header not in headers:
                headers.append(header)
    return headers


def _normalize_language_column_name(col: str) -> str:
    """Normalize language-bearing column names to include the BCP47 code."""
    m = re.match(r"^(label|hint|constraint_message|required_message|guidance_hint|image|audio|video)::\s*(.+)$", col or "")
    if not m:
        return col
    prefix, lang_part = m.groups()
    normalized = _normalize_language_tag(lang_part)
    header = normalized.get("header")
    return f"{prefix}::{header}" if header else col


def _normalize_language_columns_and_rows(
    columns: List[str],
    rows: List[Dict[str, Any]],
) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Ensure any label/hint/constraint_message language columns include the code,
    updating both the column list and row keys in place-safe copies.
    """
    all_keys = set(columns)
    for row in rows:
        all_keys.update(row.keys())

    mapping: Dict[str, str] = {}
    for key in all_keys:
        new_key = _normalize_language_column_name(key)
        mapping[key] = new_key or key

    updated_rows: List[Dict[str, Any]] = []
    for row in rows:
        new_row: Dict[str, Any] = {}
        for key, val in row.items():
            new_key = mapping.get(key, key)
            if new_key in new_row:
                # Prefer the non-empty value when a collision occurs.
                if new_row[new_key] in (None, "", " ") and val not in (None, "", " "):
                    new_row[new_key] = val
            else:
                new_row[new_key] = val
        updated_rows.append(new_row)

    updated_columns: List[str] = []
    for col in columns:
        new_col = mapping.get(col, col)
        if new_col not in updated_columns:
            updated_columns.append(new_col)

    # Include any normalized keys present in rows but missing from columns.
    for key in all_keys:
        new_key = mapping.get(key, key)
        if new_key not in updated_columns:
            updated_columns.append(new_key)

    lang_headers = _language_headers_from_columns(updated_columns)
    if lang_headers:
        # Text fields: fallback value copied to empty language columns.
        text_fields = ["label", "hint", "constraint_message", "required_message", "guidance_hint"]
        # Media fields: columns always created for each language but left empty
        # (each language may reference a different media file).
        media_fields = ["image", "audio", "video"]
        user_facing_fields = text_fields + media_fields

        col_set_now = set(updated_columns)
        for field in user_facing_fields:
            lang_cols = [f"{field}::{lang}" for lang in lang_headers]
            lang_col_set = set(lang_cols)

            # Only expand a field to language variants if the field (base column)
            # or any of its language variants is actually present in this sheet's
            # columns or row data.  This prevents, for example, hint/image/audio/video
            # columns from being auto-created on the choices sheet where they don't belong.
            field_present = field in col_set_now or any(c in col_set_now for c in lang_col_set)
            if not field_present:
                continue

            # Remove any existing lang variants (they may have been inserted in
            # non-deterministic order from the row-key collection step above),
            # then re-insert in the correct lang_headers order.
            first_idx = next(
                (i for i, c in enumerate(updated_columns) if c in lang_col_set), None
            )
            updated_columns = [c for c in updated_columns if c not in lang_col_set]
            if first_idx is not None:
                for offset, col in enumerate(lang_cols):
                    updated_columns.insert(first_idx + offset, col)
            else:
                updated_columns.extend(lang_cols)
            col_set_now = set(updated_columns)

            for row in updated_rows:
                if field in text_fields:
                    # Pick a fallback value from any existing language value or the base field.
                    fallback = None
                    for candidate_key in lang_cols + [field]:
                        val = row.get(candidate_key)
                        if val not in (None, "", " "):
                            fallback = val
                            break

                    for col in lang_cols:
                        if row.get(col) in (None, "", " ") and fallback not in (None, "", " "):
                            row[col] = fallback

                # Drop the non-language column to avoid a "default" language in ODK.
                if field in row:
                    row.pop(field, None)

            # Remove the non-language base column from the column list as well.
            updated_columns = [c for c in updated_columns if c != field]
            col_set_now = set(updated_columns)

    return updated_columns, updated_rows


def _row_has_content(values: Sequence[Any]) -> bool:
    return any(val not in (None, "", " ") for val in values)


def _infer_columns(rows: List[Dict[str, Any]], preferred: Optional[List[str]]) -> List[str]:
    seen: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.append(key)
    if preferred:
        ordered: List[str] = [col for col in preferred if col in seen]
        ordered.extend(col for col in seen if col not in ordered)
        return ordered
    return seen


def _normalize_form_spec(form_spec: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Accepts either {"sheets": {...}} or a direct mapping of sheet_name -> sheet_spec.
    Each sheet_spec may be:
      - a dict with optional "columns"/"headers" and a "rows" key, or
      - a list of row dicts (the LLM sometimes passes rows directly as the sheet value).
    """
    if "sheets" in form_spec and isinstance(form_spec["sheets"], dict):
        sheets = form_spec["sheets"]
    else:
        sheets = {k: v for k, v in form_spec.items() if isinstance(v, (dict, list))}

    normalized: Dict[str, Dict[str, Any]] = {}
    for name, sheet in sheets.items():
        if isinstance(sheet, list):
            # LLM passed a bare list of rows instead of {"columns": ..., "rows": ...}
            columns: List[str] = []
            rows: List[Dict[str, Any]] = sheet
        else:
            columns = sheet.get("columns") or sheet.get("headers") or []
            rows = sheet.get("rows") or []
        normalized[name] = {"columns": list(columns), "rows": list(rows)}
    return normalized


_LANG_COL_RE = re.compile(
    r"^(label|hint|constraint_message|required_message|guidance_hint|image|audio|video)::.+"
)


def _columns_with_data(columns: List[str], rows: List[Dict[str, Any]]) -> List[str]:
    """Return columns that have data. Language-bearing user-facing columns are always
    included even when empty so that multi-language XLSX files have a consistent
    structure (ODK requires all language columns to be present)."""
    if not rows:
        return columns
    populated: List[str] = []
    for col in columns:
        if _LANG_COL_RE.match(col):
            populated.append(col)  # always keep — may be legitimately empty
            continue
        for row in rows:
            val = row.get(col)
            if val not in (None, "", " "):
                populated.append(col)
                break
    return populated


def _write_sheet(ws, columns: List[str], rows: List[Dict[str, Any]]) -> None:
    columns = _columns_with_data(columns, rows)
    if columns:
        ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])


def _copy_sheet_values(target_wb: Workbook, source_wb, sheet_name: str) -> None:
    if sheet_name not in source_wb.sheetnames:
        return
    source_ws = source_wb[sheet_name]
    target_ws = target_wb.create_sheet(sheet_name)
    for row in source_ws.iter_rows(values_only=True):
        target_ws.append(list(row))



def design_survey_outline(
    topic: str,
    *,
    objectives: Optional[List[str]] = None,
    include_demographics: bool = True,
    languages: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    Propose a starter set of questions and choice lists for a survey.

    Returns a structure ready to merge into a form spec: survey rows and choices rows.

    When *languages* is provided (e.g. ``["English", "French"]``), the output
    uses ``label::English (en)``, ``label::French (fr)``, etc. instead of a plain
    ``label`` column so that ODK renders a multi-language form.  All labels
    are initially set to the English text as placeholders.  **The agent must
    translate every ``label::<Language (code)>`` value into the correct language
    before writing the XLSX.**

    A language-selector question (``select_one form_language``) is
    automatically prepended so that respondents can choose their preferred
    language at the start of the form.
    """
    objectives = objectives or []
    normalized_languages = _normalize_languages(languages or [])
    language_headers = [entry["header"] for entry in normalized_languages]

    # --- label / hint helpers ---
    # When multilingual, every label::<lang> is set to the English text.
    # The LLM agent is responsible for translating these before saving.
    def _label(text: str) -> Dict[str, str]:
        if not language_headers:
            return {"label": text}
        return {f"label::{lang}": text for lang in language_headers}

    def _hint(text: str) -> Dict[str, str]:
        if not language_headers:
            return {"hint": text}
        return {f"hint::{lang}": text for lang in language_headers}

    survey_rows: List[Dict[str, Any]] = []
    choices_rows: List[Dict[str, Any]] = []

    # --- language selector (only when multilingual) ---
    if normalized_languages:
        survey_rows.append(
            {
                "type": "select_one form_language",
                "name": "form_language",
                **_label("Select your preferred language"),
                "required": "yes",
            }
        )
        for entry in normalized_languages:
            lang_value = entry["code"] or _safe_form_id(entry["header"])
            choices_rows.append(
                {"list_name": "form_language", "name": lang_value, **_label(entry["name"])}
            )

    if include_demographics:
        survey_rows.extend(
            [
                {"type": "text", "name": "respondent_name", **_label("Interviewer: enter respondent name")},
                {"type": "integer", "name": "respondent_age", **_label("Respondent age"), "constraint": ". >= 0"},
                {"type": "select_one gender_list", "name": "respondent_gender", **_label("Respondent gender")},
            ]
        )
        choices_rows.extend(
            [
                {"list_name": "gender_list", "name": "female", **_label("Female")},
                {"list_name": "gender_list", "name": "male", **_label("Male")},
                {"list_name": "gender_list", "name": "other", **_label("Other / prefer not to say")},
            ]
        )

    survey_rows.extend(
        [
            {"type": "note", "name": "intro_topic", **_label(f"Survey: {topic}")},
            {
                "type": "select_one yes_no",
                "name": "consent",
                **_label("Do you consent to participate?"),
                "required": "yes",
                "constraint_message": "Consent is required to continue.",
                "relevant": "",
            },
        ]
    )

    for idx, obj in enumerate(objectives[:5], start=1):
        survey_rows.append(
            {
                "type": "text",
                "name": f"objective_{idx}",
                **_label(f"Question about: {obj}"),
                **_hint(f"Capture details related to {obj}"),
            }
        )

    survey_rows.extend(
        [
            {
                "type": "select_one frequency_list",
                "name": "usage_frequency",
                **_label("How often do you engage with this topic?"),
            },
            {
                "type": "select_one satisfaction_list",
                "name": "satisfaction_level",
                **_label("Overall satisfaction"),
                **_hint("Quick pulse check"),
            },
            {
                "type": "text",
                "name": "open_feedback",
                **_label("Any additional feedback?"),
            },
        ]
    )

    choices_rows.extend(
        [
            {"list_name": "yes_no", "name": "yes", **_label("Yes")},
            {"list_name": "yes_no", "name": "no", **_label("No")},
            {"list_name": "frequency_list", "name": "daily", **_label("Daily")},
            {"list_name": "frequency_list", "name": "weekly", **_label("Weekly")},
            {"list_name": "frequency_list", "name": "monthly", **_label("Monthly")},
            {"list_name": "frequency_list", "name": "rarely", **_label("Rarely")},
            {"list_name": "frequency_list", "name": "never", **_label("Never")},
            {"list_name": "satisfaction_list", "name": "very_satisfied", **_label("Very satisfied")},
            {"list_name": "satisfaction_list", "name": "satisfied", **_label("Satisfied")},
            {"list_name": "satisfaction_list", "name": "neutral", **_label("Neutral")},
            {"list_name": "satisfaction_list", "name": "dissatisfied", **_label("Dissatisfied")},
            {"list_name": "satisfaction_list", "name": "very_dissatisfied", **_label("Very dissatisfied")},
        ]
    )

    languages = language_headers
    return {
        "topic": topic,
        "objectives": objectives,
        "languages": languages,
        "survey_rows": survey_rows,
        "choices_rows": choices_rows,
        "notes": "Review wording, logic, and choice list names before writing XLSX."
        + (" Translations are auto-generated for common terms; review topic-specific labels." if languages else ""),
    }


def add_calculations_and_conditions(
    form_spec: Dict[str, Any],
    calculations: List[Dict[str, Any]],
    *,
    conditions: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """
    Add calculate rows and/or attach relevance/constraint logic to existing questions.
    Both `calculations` and `conditions` are optional — you may pass one or both.

    Each item in `calculations` should include:
      - name (required)
      - calculation (required)
      - label (optional but recommended)
      - relevant / constraint / constraint_message (optional)

    Each item in `conditions` should include:
      - target (question name to update)
      - relevant / constraint / constraint_message / required (any to apply)
    """
    if not calculations and not conditions:
        raise ValueError("Provide at least one calculation or condition.")

    spec = _normalize_form_spec(copy.deepcopy(form_spec))
    survey = spec.setdefault(
        "survey",
        {
            "columns": list(_PREFERRED_COLUMN_ORDER["survey"]),
            "rows": [],
        },
    )

    columns = survey.get("columns") or list(_PREFERRED_COLUMN_ORDER["survey"])
    added_calc_names: List[str] = []
    for calc in calculations:
        name = calc.get("name")
        expr = calc.get("calculation")
        if not name or not expr:
            raise ValueError("Each calculation needs both 'name' and 'calculation'.")
        row = {"type": "calculate", **calc}
        row.setdefault("label", f"Calculation for {name}")
        survey.setdefault("rows", []).append(row)
        added_calc_names.append(name)
        for key in row.keys():
            if key not in columns:
                columns.append(key)

    updated_targets: List[str] = []
    if conditions:
        targets = {r.get("name"): r for r in survey.get("rows", []) if r.get("name")}
        for cond in conditions:
            target = cond.get("target")
            if not target or target not in targets:
                continue
            row = targets[target]
            for key, val in cond.items():
                if key == "target":
                    continue
                row[key] = val
                if key not in columns:
                    columns.append(key)
            updated_targets.append(target)

    # Preserve preferred ordering but keep any new columns appended.
    preferred = _PREFERRED_COLUMN_ORDER["survey"]
    ordered = [c for c in preferred if c in columns]
    ordered.extend(c for c in columns if c not in ordered)
    survey["columns"] = ordered

    return {
        "form_spec": {"sheets": spec},
        "added_calculations": added_calc_names,
        "updated_targets": updated_targets,
        "survey_columns": survey["columns"],
    }


def merge_form_spec(
    base_spec: Dict[str, Any],
    addition_spec: Dict[str, Any],
    *,
    dedupe_by_name: bool = True,
) -> Dict[str, Any]:
    """
    Merge sheets/rows/columns from addition_spec into base_spec.

    - Preserves column order using preferred templates, then appends any new columns.
    - Optionally deduplicates rows by 'name' within each sheet.
    """
    base = _normalize_form_spec(copy.deepcopy(base_spec))
    extra = _normalize_form_spec(copy.deepcopy(addition_spec))

    summary: Dict[str, Dict[str, List[str]]] = {"added": {}, "skipped": {}}

    for sheet_name, extra_sheet in extra.items():
        base_sheet = base.setdefault(sheet_name, {"columns": [], "rows": []})
        base_rows = base_sheet.get("rows") or []
        existing_names = {r.get("name") for r in base_rows if r.get("name")}

        added: List[str] = []
        skipped: List[str] = []
        for row in extra_sheet.get("rows") or []:
            name = row.get("name")
            if dedupe_by_name and name and name in existing_names:
                skipped.append(name)
                continue
            base_rows.append(row)
            if name:
                existing_names.add(name)
                added.append(name)
        base_sheet["rows"] = base_rows

        columns = base_sheet.get("columns") or []
        col_set = set(columns)
        for col in extra_sheet.get("columns") or []:
            if col not in col_set:
                columns.append(col)
                col_set.add(col)

        preferred = _PREFERRED_COLUMN_ORDER.get(sheet_name.lower())
        if preferred:
            ordered = [c for c in preferred if c in col_set]
            ordered.extend(c for c in columns if c not in ordered)
            columns = ordered
        base_sheet["columns"] = columns

        summary["added"][sheet_name] = added
        summary["skipped"][sheet_name] = skipped

    return {
        "merged_spec": {"sheets": base},
        "summary": summary,
    }


def new_form_spec(
    form_title: str,
    *,
    form_id: Optional[str] = None,
    default_language: str = "English",
    languages: Optional[List[str]] = None,
    version: Optional[str] = None,
    submission_url: Optional[str] = None,
    public_key: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Build a starter form spec with blank survey/choices sheets and basic settings.

    When *languages* is provided (e.g. ``["English", "French"]``), the survey
    and choices column lists include ``label::<Language (code)>`` columns for each
    language (e.g., ``label::English (en)``), and the settings ``default_language``
    is set to the first language.  A language-selector question is added to the
    survey sheet.

    The returned structure matches what `write_xlsform(...)` expects.
    """
    form_id = form_id or _safe_form_id(form_title)
    version = version or datetime.datetime.now().strftime("%Y%m%d%H%M")
    normalized_languages = _normalize_languages(languages or [])
    language_headers = [entry["header"] for entry in normalized_languages]

    # Build column lists — only include label::<lang> when languages are given.
    survey_cols = list(_PREFERRED_COLUMN_ORDER["survey"])
    choices_cols = list(_PREFERRED_COLUMN_ORDER["choices"])
    if language_headers:
        default_language = language_headers[0]
        # Insert label::<lang> columns right after "label" (or replace it)
        for col_list in (survey_cols, choices_cols):
            idx = col_list.index("label")
            lang_cols = [f"label::{lang}" for lang in language_headers]
            col_list[idx:idx + 1] = lang_cols

    survey_rows: List[Dict[str, Any]] = []
    choices_rows: List[Dict[str, Any]] = []

    # Add language selector when multilingual
    if language_headers:
        def _label(text: str) -> Dict[str, str]:
            return {f"label::{lang}": text for lang in language_headers}

        survey_rows.append(
            {
                "type": "select_one form_language",
                "name": "form_language",
                **_label("Select your preferred language"),
                "required": "yes",
            }
        )
        for entry in normalized_languages:
            lang_value = entry["code"] or _safe_form_id(entry["header"])
            choices_rows.append(
                {"list_name": "form_language", "name": lang_value, **_label(entry["name"])}
            )

    return {
        "sheets": {
            "survey": {
                "columns": survey_cols,
                "rows": survey_rows,
            },
            "choices": {
                "columns": choices_cols,
                "rows": choices_rows,
            },
            "settings": {
                "columns": _PREFERRED_COLUMN_ORDER["settings"],
                "rows": [
                    {
                        "form_title": form_title,
                        "form_id": form_id,
                        "default_language": default_language,
                        "version": version,
                        "submission_url": submission_url or "",
                        "public_key": public_key or "",
                    }
                ],
            },
        }
    }


def load_xlsform(
    file_path: str,
    *,
    sheet_names: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    Read an existing XLSForm and return a structured, JSON-friendly spec.
    Skips fully empty rows to keep the payload concise.
    """
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"XLSForm not found: {path}")

    wb = load_workbook(path, data_only=True)
    selected = sheet_names or wb.sheetnames

    sheets: Dict[str, Dict[str, Any]] = {}
    for sheet_name in selected:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[sheet_name] = {"columns": [], "rows": []}
            continue
        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header = [col for col in header if col]  # drop blanks in the header row

        parsed_rows: List[Dict[str, Any]] = []
        for raw in rows[1:]:
            if not _row_has_content(raw):
                continue
            row_dict = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
            parsed_rows.append(row_dict)

        sheets[sheet_name] = {"columns": header, "rows": parsed_rows}

    return {
        "path": str(path),
        "sheets": sheets,
        "sheet_names": list(sheets.keys()),
        "row_counts": {k: len(v["rows"]) for k, v in sheets.items()},
    }


def write_xlsform(
    form_spec: Dict[str, Any],
    output_path: Optional[str] = None,
    *,
    base_form_path: Optional[str] = None,
    preserve_additional_sheets: bool = True,
) -> Dict[str, Any]:
    """
    Save a structured form_spec to an XLSX file.

    When no base_form_path is given the official ODK XLSForm Template is
    downloaded (and cached) so that every generated file inherits the
    template's structure and formatting.

    The form_spec can be either {"sheets": {...}} or a mapping of sheet_name -> sheet_spec.
    Each sheet_spec should contain "rows" (list of dicts) and optional "columns".
    """
    spec = _normalize_form_spec(form_spec)
    if not spec:
        raise ValueError("form_spec is empty; include at least the survey sheet.")

    preferred_copy: Dict[str, List[str]] = {
        name.lower(): order for name, order in _PREFERRED_COLUMN_ORDER.items()
    }

    # Determine the base workbook: an explicit path, the official ODK
    # template (downloaded & cached), or – as last resort – a blank workbook.
    if base_form_path:
        base_path = Path(base_form_path).expanduser().resolve()
        if not base_path.exists():
            raise FileNotFoundError(f"Base form not found: {base_path}")
    else:
        base_path = _get_template_path()

    if base_path and base_path.exists():
        wb = load_workbook(base_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    def _prepare_sheet(sheet_name: str, sheet_spec: Dict[str, Any]) -> tuple[List[str], List[Dict[str, Any]]]:
        columns = sheet_spec.get("columns") or []
        rows = sheet_spec.get("rows") or []
        columns = columns or _infer_columns(rows, preferred_copy.get(sheet_name.lower()))
        if not columns:
            columns = ["type", "name", "label"] if sheet_name.lower() == "survey" else ["name", "label"]
        # Normalize language columns first so language-variant columns (label::lang)
        # are created before we apply preferred ordering.
        columns, rows = _normalize_language_columns_and_rows(columns, rows)
        preferred = preferred_copy.get(sheet_name.lower())
        if preferred:
            col_set = set(columns)
            placed: set[str] = set()
            ordered: List[str] = []
            for pref_col in preferred:
                if pref_col in col_set and pref_col not in placed:
                    ordered.append(pref_col)
                    placed.add(pref_col)
                # Expand base column to its language variants in order
                lang_variants = [c for c in columns if c.startswith(f"{pref_col}::") and c not in placed]
                for lv in lang_variants:
                    ordered.append(lv)
                    placed.add(lv)
            # Append anything not already placed (unknown/extra columns)
            for col in columns:
                if col not in placed:
                    ordered.append(col)
            columns = ordered
        return columns, rows

    normalized_sheets: Dict[str, tuple[List[str], List[Dict[str, Any]]]] = {}
    survey_languages: List[str] = []
    for sheet_name, sheet_spec in spec.items():
        if sheet_name.lower() == "settings":
            continue
        columns, rows = _prepare_sheet(sheet_name, sheet_spec)
        if sheet_name.lower() == "survey":
            survey_languages = _language_headers_from_columns(columns)
        normalized_sheets[sheet_name] = (columns, rows)

    if "settings" in spec:
        settings_spec = spec["settings"]
        settings_rows = settings_spec.get("rows") or []
        if survey_languages:
            for row in settings_rows:
                if not row.get("default_language"):
                    row["default_language"] = survey_languages[0]
        settings_columns = settings_spec.get("columns") or _PREFERRED_COLUMN_ORDER["settings"]
        settings_columns, settings_rows = _normalize_language_columns_and_rows(settings_columns, settings_rows)
        normalized_sheets["settings"] = (settings_columns, settings_rows)

    written: List[str] = []
    for sheet_name, (columns, rows) in normalized_sheets.items():
        # Reuse the existing template sheet (clearing its rows) or create new.
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        _write_sheet(ws, columns, rows)
        written.append(sheet_name)

    if not preserve_additional_sheets:
        for name in list(wb.sheetnames):
            if name not in written:
                del wb[name]

    default_output = "odk_form_draft.xlsx"
    if base_form_path:
        bp = Path(base_form_path).expanduser().resolve()
        default_output = bp.with_name(bp.stem + "_draft.xlsx").name
    output = Path(output_path or default_output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "output_path": str(output),
        "sheet_names": wb.sheetnames,
        "row_counts": {name: len(spec.get(name, {}).get("rows", [])) for name in written},
    }

def save_xlsform_draft(
    form_spec: Dict[str, Any],
    output_path: Optional[str] = None,
    *,
    base_form_path: Optional[str] = None,
    preserve_additional_sheets: bool = True,
) -> Dict[str, Any]:
    """
    Convenience wrapper that always emits an XLSX, using a timestamped filename when none is provided.
    Automatically runs pyxform validation after writing and includes the result.
    """
    if not output_path:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = "odk_form"
        if base_form_path:
            stem = Path(base_form_path).stem + "_draft"
        output_path = f"{stem}_{ts}.xlsx"
    result = write_xlsform(
        form_spec,
        output_path=output_path,
        base_form_path=base_form_path,
        preserve_additional_sheets=preserve_additional_sheets,
    )
    result["validation"] = validate_xlsform(result["output_path"])
    return result


def validate_xlsform(file_path: str) -> Dict[str, Any]:
    """
    Validate an XLSForm XLSX file using pyxform.

    Converts the XLSX to XForm (XML) in memory and captures any errors or warnings.
    Returns {"valid": bool, "errors": [...], "warnings": [...]}.
    If errors is non-empty the form is invalid and must be fixed before use.
    """
    try:
        from pyxform.xls2xform import xls2xform_convert
    except ImportError:
        return {"valid": None, "errors": ["pyxform is not installed"], "warnings": []}

    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"XLSForm not found: {path}")

    tmp_xform = Path(tempfile.mktemp(suffix=".xml"))
    try:
        warnings = xls2xform_convert(
            xlsform_path=str(path),
            xform_path=str(tmp_xform),
            validate=False,   # skip Java ODK Validate; pyxform itself catches structural errors
            pretty_print=False,
        )
        return {
            "valid": True,
            "errors": [],
            "warnings": [str(w) for w in (warnings or [])],
        }
    except Exception as exc:
        # pyxform raises PyXFormError (or subclasses) for structural mistakes
        error_text = str(exc)
        # Split into individual lines for readability
        errors = [line.strip() for line in error_text.splitlines() if line.strip()]
        return {
            "valid": False,
            "errors": errors,
            "warnings": [],
        }
    finally:
        tmp_xform.unlink(missing_ok=True)



INSTRUCTION = """
You are odk_xlsform_agent. You help users design or edit ODK XLSForms by following a strict step-by-step procedure. Complete each step fully and confirm with the user before moving on. Never skip ahead.

---

## STEP-BY-STEP PROCEDURE

### STEP 1 — New form or existing form?
Ask the user: "Do you want to create a new form, or load and edit an existing one?"

**If editing an existing form:**
- Call `load_xlsform(file_path)` with the path the user provides.
- Summarize what you loaded: sheet names, row counts, column names, and any choice lists found.
- Ask the user to confirm that you have loaded the right file and that your summary is accurate.
- ⏸ Wait here — the user must confirm the correct file was loaded before you make any changes.

**If creating a new form:**
- Proceed to STEP 2.

---

### STEP 2 — Language support
Ask the user: "Should this form support multiple languages, or just one? If multiple, which languages?"

**If single language:**
- Note the language (default to English if unspecified). No special column handling needed.
- Proceed immediately to STEP 3.

**If multiple languages:**
- List the languages back to the user with their IANA codes (e.g. English → en, French → fr, Arabic → ar, Chinese → zh, Swahili → sw).
- Explain: the first language listed will be the default; a language-selector question will be added automatically at the start of the form.
- Confirm the list and ordering with the user, then proceed to STEP 3 immediately unless they want changes.

**Column format rule (apply throughout all subsequent steps):**
- Use `label::LanguageName(code)` — e.g. `label::English(en)`, `label::French(fr)`. No space before the parenthesis.
- All user-facing columns must have a variant per language: `label`, `hint`, `constraint_message`, `required_message`, `guidance_hint`, `image`, `audio`, `video`.

---

### STEP 3 — Scaffold the form
Call `new_form_spec(form_title, languages=[...])` using the title and languages confirmed in previous steps.

- Show the user the resulting settings (form_title, form_id, default_language, version) and the starter column structure.
- Proceed immediately to STEP 4. Mention the settings briefly and note that the user can ask to change them at any time.

---

### STEP 4 — Design the questions
First, assess whether the user has already provided specific question details:

**If the user already provided a detailed spec** (specific questions, sections, field types, choice lists, skip logic, calculations — as in a structured message or loaded document):
- Do NOT call `design_survey_outline`. Do NOT ask "auto or manual?".
- Read the spec directly and construct the full survey rows and choice rows yourself.
- Translate the user's spec into proper XLSForm rows: assign types (text, integer, select_one, select_multiple, note, calculate, begin_group, end_group, begin_repeat, end_repeat, etc.), names (snake_case), labels per language, required flags, and choice list names.
- For any choice lists mentioned, build the corresponding choices rows.
- Display the complete proposed question table to the user, organised by section.
- ⏸ Wait here — the user must approve the question list before you merge and build on it.

**If the user has NOT provided specific questions** (they only gave a topic or vague goal):
- Offer two options:
  a) Auto-generate a starter outline with `design_survey_outline(topic, objectives=[...], languages=[...])` — best when the user has no specific questions in mind and just wants a starting point.
  b) Build the question list manually, gathering questions section by section.
- `design_survey_outline` produces a generic template (demographics, consent, frequency, satisfaction, open feedback). Only use it when the user explicitly wants this kind of generic structure. Never use it when the user has described specific questions.
- Display the proposed rows in a readable table and ask: "Does this look right? What should we add, remove, or change?"
- ⏸ Wait here — the user must approve the question list before you merge and build on it.

---

### STEP 5 — Merge the outline into the form spec
Call `merge_form_spec(base_spec, addition_spec)` to fold the confirmed questions and choice rows into the scaffolded form spec.

- Show a summary: how many rows were added to each sheet, and how many were skipped as duplicates.
- Proceed immediately to STEP 6 without waiting — the merge is mechanical and the user can flag issues at the next review.

---

### STEP 6 — Skip logic, constraints, and calculations
Based on the form's topic, the description document (if loaded), and the confirmed question list, **proactively identify** any skip logic, validation rules, or derived fields that would make sense. Do not ask the user whether they want these — reason about the form yourself and present your findings.

For each item you identify, explain it clearly to the user. For example:
- "Because you have a consent question, I suggest making all subsequent questions relevant only when the consent answer is 'yes' (XPath: $consent = 'yes')."
- "Since you ask for respondent age, I suggest a `calculate` row to derive an age group (child / adult / elder) for use in skip logic later."
- "The age field should have a constraint `. >= 0 and . <= 120` to catch data entry errors."

Present all identified conditions and calculations as a structured list, showing for each:
- **Type**: relevant condition / constraint / calculation
- **Target question** (or new calculate row name)
- **Proposed XPath expression**
- **Reason**: one sentence explaining why this makes sense given the form's purpose

Then ask the user: "Here are the conditions and calculations I identified. Do you agree with these? Would you like to add, remove, or change any?"

**After the user confirms or adjusts the list:**
- Call `add_calculations_and_conditions(form_spec, calculations=[...], conditions=[...])` with the agreed items.
- Display the updated survey rows and affected columns.
- ⏸ Wait here — the user must confirm the logic is correct before you proceed.

**If no conditions or calculations are warranted** (e.g. the form is very simple with no logic dependencies):
- State this explicitly: "I did not find any skip logic or derived fields needed for this form." Then proceed directly to STEP 7.

---

### STEP 7 — Translation verification (multilingual forms only; skip if single language)
This step is MANDATORY before saving a multilingual form. Do not skip it.

1. Identify the **primary language** (the first language the user specified).
2. Scan **every row** in the survey sheet and the choices sheet. For each non-primary-language column (e.g. `label::French(fr)`, `hint::Arabic(ar)`):
   - If the value is **empty** and the primary-language value is non-empty → translate and fill it in.
   - If the value is **identical to the primary-language value** → it is an untranslated placeholder → translate and replace it.
   - Leave a cell empty only if the primary-language cell is also empty (e.g. media columns with no file).
3. Present a summary table to the user showing which rows/columns were translated and which were intentionally left empty.
4. Ask the user: "Do these translations look correct? Shall I proceed to save?"
- ⏸ Wait here — the user must approve the translations before you write the file.

You have strong multilingual capabilities — produce accurate, natural translations, not word-for-word copies.

---

### STEP 8 — Save, validate, and export
Once the user confirms everything is ready:

1. Call `save_xlsform_draft(form_spec, output_path)`. Use a timestamped filename if the user does not specify one.
   - The result includes a `"validation"` key automatically: `{"valid": bool, "errors": [...], "warnings": [...]}`.

2. **Check the validation result immediately — do not show the file to the user yet.**
   - If `valid` is `False`: the form has structural errors that will prevent it from loading in ODK.
     - Read each error message carefully.
     - Fix the identified problems in the form spec (wrong type names, broken XPath expressions, missing choice lists, unbalanced groups, etc.).
     - Call `save_xlsform_draft` again with the corrected spec.
     - Repeat until `valid` is `True`. Do not present the file to the user while errors remain.
   - If `valid` is `True` but `warnings` is non-empty: show the warnings to the user and explain what they mean. Warnings do not block the form from loading but may indicate sub-optimal logic.

3. Once validation passes, report to the user:
   - The output file path.
   - Sheet names and row counts.
   - Any warnings (if present).
   - Offer to make further edits or save another version.

---

## GENERAL RULES (apply throughout all steps)

**Never expose internal step labels:**
Do NOT say "STEP 1", "STEP 5", "Step 7", "moving to step", "proceeding to STEP", or any similar phrasing in your replies to the user. Follow the procedure internally, but communicate naturally and conversationally — the user should never see step numbers.

**Proceed autonomously whenever you can:**
If the user's message already contains enough information to complete one or more steps, execute those steps immediately and show the results — do not stop to ask for permission to proceed. Chain as many steps as the available information allows in a single response. Only pause and ask the user when you have reached a genuine decision point: missing information you cannot infer, a significant design choice the user should own, or a step whose output the user must review before you can continue meaningfully.

**When to pause and wait for the user:**
- You are missing a required piece of information (e.g. the form title, the language list, a file path) that you cannot reasonably infer.
- You have produced a full question list or a set of conditions/calculations and need the user to verify the design is correct before building on it further.
- Translations are complete and ready for user review before saving.
- You are about to write the final XLSX file.

**When NOT to pause:**
- Moving from scaffolding straight into question design when the topic is already clear.
- Running `merge_form_spec` immediately after the question list is approved, without waiting for an explicit "merge now" command.
- Skipping the skip-logic step when no logic is warranted, and stating your reasoning, rather than asking whether the user wants it.

**Column conventions:**
- XLSForm column order: type → name → label (or label::Lang) → hint → required → relevant → appearance → default → constraint → constraint_message → calculation → choice_filter → repeat_count.
- For select_one/select_multiple questions, always ensure the matching choice list exists in the choices sheet and the list name is spelled consistently.
- Support groups and repeats via `begin group`/`end group` and `begin repeat`/`end repeat` rows (note: space, not underscore); always keep grouping balanced (every begin has a matching end).

---

## XLSFORM SYNTAX REFERENCE (MANDATORY — always follow these rules exactly)

NOTATION NOTE: In all examples below, variable references are written as $[field_name] using square brackets.
When generating actual XLSForm content, ALWAYS convert this notation by replacing the square brackets with curly braces:
- The opening square bracket [ becomes an opening curly brace
- The closing square bracket ] becomes a closing curly brace
- Example: $[age] in these docs → dollar sign + curly-open + age + curly-close in the real form
This applies to every variable reference in every expression column: relevant, constraint, calculation, choice_filter, repeat_count, and default.

### Question types
- `text` — free text
- `integer` — whole numbers only
- `decimal` — decimal numbers
- `date` / `time` / `dateTime` — date/time pickers
- `select_one list_name` — single-choice (radio); always a space before the list name
- `select_multiple list_name` — multi-choice (checkbox); always a space before the list name
- `note` — read-only display text; no user input
- `calculate` — computed field; write the expression in the `calculation` column, not the `label` column
- `hidden` — invisible field; stores a computed or pre-populated value
- `acknowledge` — user must tap to confirm before proceeding
- `begin group` / `end group` — group rows; use a space, not underscore
- `begin repeat` / `end repeat` — repeat group rows; use a space, not underscore

### Referencing other questions
- Always use a dollar sign followed by the field name in curly braces to reference another field's value in any expression column. Example: to reference a field named "age", write $[age] (in real XPath this is the dollar-sign curly-brace form).
- Inside a `constraint` column only, use `.` (a single dot) to refer to the current field's own value — never use the dollar-sign reference to the same field inside its own constraint.
- Inside a repeat, use `position(..)` (1-based integer) to get the current repeat instance index.

### Operators
- Math: `+`, `-`, `*`, `div` (not `/`), `mod`
- Comparison: `=`, `!=`, `>`, `>=`, `<`, `<=`
- Boolean: `and`, `or`, `not(...)`
- String concatenation: use `concat(a, b, ...)` — the `+` operator does NOT concatenate strings.

### Essential functions
| Function | Correct usage (using $[field] notation for variable refs) |
|---|---|
| `if(expr, then, else)` | `if($[age] >= 18, 'adult', 'minor')` |
| `coalesce(a, b)` | Returns `a` if non-empty, else `b`; use to guard empty numerics |
| `selected(field, 'value')` | Test if a select_multiple answer includes 'value': `selected($[tools], 'chatgpt')` |
| `count-selected(field)` | Count selected options in a select_multiple: `count-selected($[tools])` |
| `concat(a, b, ...)` | String concatenation: `concat($[first_name], ' ', $[last_name])` |
| `round(number, places)` | `round($[score] * 1.18, 2)` |
| `int(number)` | Truncate to integer (does not round): `int($[score])` |
| `string(arg)` | Convert to string |
| `not(expr)` | Boolean negation: `not(selected($[tools], 'none'))` |
| `regex(., 'pattern')` | In a constraint, test current value against a full regex match |
| `contains(str, sub)` | True if sub found inside str |
| `string-length(str)` | Character count; use in constraints: `string-length(.) <= 500` |
| `count(nodeset)` | Count repeat instances: `count($[people])` |
| `sum(nodeset)` | Sum a field across repeat instances |
| `indexed-repeat(field, group, i)` | Access repeat instance i's value of field |
| `today()` | Current date |
| `now()` | Current datetime (re-evaluates on every change — use `once(now())` or `trigger` to stabilize) |
| `uuid()` | RFC 4122 v4 UUID |

### Relevant (skip logic)
- Column: `relevant`
- Returns `True` to show the question/group, `False` to hide it.
- Correct examples (using $[field] notation — in real XPath use dollar-curly-brace form):
  - `$[consent] = 'yes'`
  - `selected($[tools], 'chatgpt')`
  - `$[age] >= 18 and $[age] <= 65`
  - `$[frequency] = 'daily' or $[frequency] = 'weekly'`
- Apply `relevant` to a `begin group` row to show/hide an entire section at once.

### Constraints
- Column: `constraint`
- Returns `True` to accept the answer, `False` to reject it (showing `constraint_message`).
- Use `.` for the current field's value inside its own constraint.
- Correct examples:
  - `. >= 0 and . <= 120` (age range)
  - `. >= 16 and . <= 80`
  - `count-selected(.) >= 1` (at least one choice selected)
  - `string-length(.) <= 500` (max character limit)
  - `regex(., '[0-9]+')` (digits only)
- Constraints do NOT fire on blank answers — always combine with `required: yes` to enforce non-blank.

### Calculations
- Column: `calculation` (only used when `type` is `calculate`)
- Empty/unanswered number fields produce an empty string, NOT 0. Always guard with coalesce or if:
  - BAD: `$[score_a] + $[score_b]` (crashes if either is empty)
  - GOOD: `coalesce($[score_a], 0) + coalesce($[score_b], 0)`
- Assign numeric scores with nested if():
  - `if($[frequency] = 'daily', 4, if($[frequency] = 'weekly', 3, if($[frequency] = 'monthly', 2, 1)))`
- Profile classification with nested if():
  - `if($[final_score] > 10, 'High Adoption', if($[final_score] >= 5, 'Moderate Adoption', 'Low Adoption'))`
- Count selected options: `count-selected($[tools])`
- Multiply two calculated fields: `$[ai_usage_score] * $[tool_count]`

### Choice filter (cascading selects)
- Column: `choice_filter`
- The expression references **column names in the choices sheet**, not question names.
- Example: choices sheet has a `region` column; filter by: `region = $[selected_region]`

### Repeat groups
- `repeat_count` can be a fixed number or a reference to an integer answer field.
- Access a specific iteration from outside the repeat: `indexed-repeat($[child_name], $[children], 2)`
- `position(..)` inside a repeat returns the current 1-based iteration number.

### Common mistakes to avoid
1. Using `/` for division — use `div` instead: `$[a] div $[b]`
2. Using `+` to join strings — use `concat(...)` instead
3. Referencing the current field by name inside its own `constraint` — use `.` instead
4. Forgetting `coalesce()` around numeric fields that may be empty before arithmetic
5. Using `begin_group` / `end_group` with underscores — correct spelling is `begin group` / `end group` (space)
6. Writing `select_one` without a space before the list name — correct: `select_one my_list`
7. Referencing a `select_multiple` field with `= 'value'` — use `selected($[field], 'value')` instead
8. Using `now()` in a calculation that should only evaluate once — wrap with `once(now())` or use the `trigger` column

**When in doubt:**
- Ask clarifying questions instead of guessing.
- If the user's request is ambiguous, present two or three concrete options and let them choose.
""".strip()

root_agent = Agent(
    name="odk_xlsform_agent",
    # model=LiteLlm(model="ollama/llama3.1:8b"),
    model="gemini-2.5-flash",
    instruction=INSTRUCTION,
    tools=[
        design_survey_outline,
        merge_form_spec,
        add_calculations_and_conditions,
        new_form_spec,
        load_xlsform,
        write_xlsform,
        save_xlsform_draft,
        validate_xlsform,
    ],
)
